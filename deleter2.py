import openpyxl
from openpyxl.styles import PatternFill

# Defining file paths
# admin_file = '/Users/aloncohen/Desktop/admin test.xlsx'
admin_file = '#insert file path here'

wb_obj = openpyxl.load_workbook(admin_file)
sheet1 = wb_obj['COVID19_ZH_V5_Total']

delete_list = []

while True:
    answer = input('Enter ID to be removed, confirm with Enter, to stop enter stop: ')
    if answer == 'stop':
        break
    delete_list.append(answer)

counter = 11
for row in sheet1:

    ID = str(sheet1['A' + str(counter)].value)

    try:
        termin = sheet1['E' + str(counter)].value
    except:
        AttributeError

    if ID in delete_list:
        sheet1['G' + str(counter)] = 'Termin verpasst'
        sheet1['H' + str(counter)] = None
        sheet1['J' + str(counter)] = None
        sheet1['K' + str(counter)] = None
        sheet1['L' + str(counter)] = None
        sheet1['L' + str(counter)].fill = PatternFill(fgColor="FFF2CC", fill_type="solid")
        sheet1['M' + str(counter)] = None
        sheet1['M' + str(counter)].fill = PatternFill(fgColor="FCE4D6", fill_type="solid")
        print('deleted: ' + ID)

    counter += 1

wb_obj.save('#insert file path here')
#wb_obj.save('/Users/aloncohen/Desktop/admin test.xlsx')

print(delete_list)