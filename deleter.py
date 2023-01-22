import openpyxl
from openpyxl.styles import PatternFill

#admin_file = 'admintestfile.xlsx'
admin_file = '#insert file path here'

wb_obj = openpyxl.load_workbook(admin_file)
sheet1 = wb_obj['COVID19_ZH_V5_Total']

delete_list = []

while True:
    answer = input('Enter ID to be removed, confirm with Enter, to stop enter stop: ')
    if answer == 'stop':
        break
    delete_list.append(answer)

counter = 11

for row in sheet1:

    ID = str(sheet1['D' + str(counter)].value)
    redCapFallErstellt = sheet1['B' + str(counter)].value

    try:
        termin = sheet1['E' + str(counter)].value
    except:
        AttributeError

    if ID in delete_list:
        if redCapFallErstellt == 1:
            sheet1['B' + str(counter)] = 'RedCap erfasst, Termin verpasst'
            sheet1['C' + str(counter)] = None
            sheet1['E' + str(counter)] = None
            sheet1['F' + str(counter)] = None
            sheet1['H' + str(counter)] = None
            sheet1['H' + str(counter)].fill = PatternFill(fgColor="DDEBF7", fill_type="solid")
            sheet1['I' + str(counter)] = None
            sheet1['I' + str(counter)].fill = PatternFill(fgColor="FFF2CC", fill_type="solid")
            sheet1['J' + str(counter)] = None
            sheet1['J' + str(counter)].fill = PatternFill(fgColor="FCE4D6", fill_type="solid")
            sheet1['L' + str(counter)] = None
            print('deleted: ' + ID)
    counter += 1

wb_obj.save('#insert file path here')
#wb_obj.save('admintestfile.xlsx')

print(delete_list)
