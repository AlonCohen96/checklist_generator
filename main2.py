# pip install openpyxl
# pip install python-docx
# pip install docxtpl

import openpyxl
from openpyxl.styles import PatternFill
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm
import datetime

# Defining file paths
admin_file = '' #insert file path here
#admin_file = '/Users/aloncohen/Desktop/admin test.xlsx'

# Use this to save the excel with formulas
wb_obj = openpyxl.load_workbook(admin_file)
sheet1 = wb_obj['COVID19_ZH_V5_Total']
sheet2 = wb_obj['counter']

# Use these to remove formulas (read-only)
wb_obj_f = openpyxl.load_workbook(admin_file, data_only=True)
sheet1_f = wb_obj_f['COVID19_ZH_V5_Total']
sheet2_f = wb_obj_f['counter']

# Opening label template
wb_label_template = openpyxl.load_workbook('dependencies/labels_template_phase5.xlsx')
label_sheet = wb_label_template.active

# Accessing the template checklist file
doc = DocxTemplate('dependencies/checklist_template_juni.docx')

# Defining group values
total_group2 = sheet1_f['L3'].value

# Defining some static strings and images to be inserted into the checklist later on
header_g2 = InlineImage(doc, 'dependencies/g2_header.png', width=Mm(170))
header_g3 = InlineImage(doc, 'dependencies/g3_header.png', width=Mm(170))

small_g2 = InlineImage(doc, 'dependencies/g2_small.png', height=Mm(5))
small_g3 = InlineImage(doc, 'dependencies/g3_small.png', height=Mm(5))

tubes_g2 = '□ 1x Serum (8.5ml)\n' \
           '□ 1x Serum (8.5ml)\n' \
           '□ 1x Heparin (4ml)'

tubes_g3 = '□ 1x Serum (8.5ml)\n' \
           '□ 1x Serum (8.5ml)'

n_serum_g2_g3 = '2x'

heparin_g1_g2 = '□ Handling Heparin (4ml)\n' \
                '\t\t1 T-Cell reactivity test'

# Counts how many participants of group 2 are scheduled for each test day (max 24 per day)

test_days_counter = {'07.06.2022': ['I1', sheet2_f['J1'].value],
                     '08.06.2022': ['I2', sheet2_f['J2'].value],
                     '09.06.2022': ['I3', sheet2_f['J3'].value],
                     '10.06.2022': ['I4', sheet2_f['J4'].value],
                     '11.06.2022': ['I5', sheet2_f['J5'].value],

                     '13.06.2022': ['I7', sheet2_f['J7'].value],
                     '14.06.2022': ['I8', sheet2_f['J8'].value],
                     '15.06.2022': ['I9', sheet2_f['J9'].value],
                     '16.06.2022': ['I10', sheet2_f['J10'].value],
                     '17.06.2022': ['I11', sheet2_f['J11'].value],
                     '18.06.2022': ['I12', sheet2_f['J12'].value],

                     '20.06.2022': ['I14', sheet2_f['J14'].value],
                     '21.06.2022': ['I15', sheet2_f['J15'].value],
                     '22.06.2022': ['I16', sheet2_f['J16'].value],
                     '23.06.2022': ['I17', sheet2_f['J17'].value],
                     '24.06.2022': ['I18', sheet2_f['J18'].value],
                     '25.06.2022': ['I19', sheet2_f['J19'].value],

                     '27.06.2022': ['I21', sheet2_f['J21'].value],
                     '28.06.2022': ['I22', sheet2_f['J22'].value],
                     '29.06.2022': ['I23', sheet2_f['J23'].value],
                     '30.06.2022': ['I24', sheet2_f['J24'].value],
                     '01.07.2022': ['I25', sheet2_f['J25'].value],
                     
                     '04.07.2022': ['I28', sheet2_f['J28'].value],
                     '05.07.2022': ['I29', sheet2_f['J29'].value],
                     '06.07.2022': ['I30', sheet2_f['J30'].value],
                     '07.07.2022': ['I31', sheet2_f['J31'].value],
                     '08.07.2022': ['I32', sheet2_f['J32'].value]}

# Allocates a participant to group 1, 2, or 3 according to criteria
def group_allocater():

    global total_group2
    global counter

    # Saturdays automatically Group 3
    if termin == '11.06.2022' or termin == '18.06.2022' or termin == '25.06.2022' or termin == '28.06.2022' or termin == '05.07.2022':
        sheet1['M' + str(counter)].value = 1
        sheet1['M' + str(counter)].fill = PatternFill(fgColor="FCE4D6", fill_type="solid")
        label_generator_g3()

        return [small_g3, tubes_g3, n_serum_g2_g3, '', '', header_g3]

    else:
        if oldGroup2 == 1 and total_group2 < 199 and test_days_counter[termin][1] < 24:
            # Group 2
            test_days_counter[termin][1] += 1
            total_group2 += 1
            sheet2[test_days_counter[termin][0]].value += 1
            sheet1['L' + str(counter)].value = 1
            sheet1['L' + str(counter)].fill = PatternFill(fgColor="FFF2CC", fill_type="solid")
            label_generator_g2()

            return [small_g2, tubes_g2, n_serum_g2_g3, '', heparin_g1_g2, header_g2]

        else:
            # Group 3
            sheet1['M' + str(counter)].value = 1
            sheet1['M' + str(counter)].fill = PatternFill(fgColor="FCE4D6", fill_type="solid")
            label_generator_g3()

            return [small_g3, tubes_g3, n_serum_g2_g3, '', '', header_g3]


# Generating labels for group 2
def label_generator_g2():

    global lc

    label_ID = 'CI-T2-' + str(ID) + '-' + birthdate[-2:]

    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = ''
    label_sheet['C' + str(lc)] = termin
    lc += 1
    for i in range(3):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'BLOOD'
        label_sheet['C' + str(lc)] = termin
        lc += 1
    for i in range(6):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'SERUM'
        label_sheet['C' + str(lc)] = termin
        lc += 1


# Generating labels for group 3
def label_generator_g3():

    global lc

    label_ID = 'CI-T2-' + str(ID) + '-' + birthdate[-2:]

    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = ''
    label_sheet['C' + str(lc)] = termin
    lc += 1
    for i in range(1):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'BLOOD'
        label_sheet['C' + str(lc)] = termin
        lc += 1
    for i in range(1):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'SERUM'
        label_sheet['C' + str(lc)] = termin
        lc += 1

# Iterating through the admin file. lc stands for label counter
counter = 11
lc = 2
for row in sheet1:

    # Extracting the data needed cell by cell
    ID = sheet1['A' + str(counter)].value
    vorname = sheet1['P' + str(counter)].value
    name = sheet1['Q' + str(counter)].value
    sex = sheet1['U' + str(counter)].value
    if sex == 1:
        sex = 'male'
    else:
        sex = 'female'

    try:
        birthdate = datetime.datetime.strptime(str(sheet1['R'+str(counter)].value.date()), "%Y-%m-%d").strftime("%d.%m.%Y")
    except AttributeError:
        birthdate = 'n.a.'

    try:
        termin = datetime.datetime.strptime(str(sheet1['G' + str(counter)].value.date()), "%Y-%m-%d").strftime("%d.%m.%Y")
    except AttributeError:
        termin = 'n.a.'

    try:
        time = sheet1['H' + str(counter)].value.strftime("%H:%M")
    except AttributeError:
        time = 'n.a.'

    fragebogenCode = sheet1['I' + str(counter)].value
    studienMappeGeneriert = sheet1['J' + str(counter)].value
    oldGroup2 = sheet1['E' + str(counter)].value

    # Pick tomorrow's appointments only or Mondays as well
    if termin == (datetime.date.today() + datetime.timedelta(days=1)).strftime("%d.%m.%Y") or termin == (datetime.date.today() + datetime.timedelta(days=3)).strftime("%d.%m.%Y"):

    # Pick today's appointments only
    #if termin == datetime.date.today().strftime("%d.%m.%Y"):

        # Pick relevant participants and generate the individual documents for them, update the admin file
        if studienMappeGeneriert is None:
            labsheet_inserts = group_allocater()
            context = {'ID': ID,
                       'vorname': vorname,
                       'name': name,
                       'sex': sex,
                       'birthdate': birthdate,
                       'termin': termin,
                       'time': time,
                       'fragebogenCode': fragebogenCode,
                       'label_ID': 'CI-T2-' + str(ID) + '-' + birthdate[-2:],
                       'group': labsheet_inserts[0],
                       'tubes': labsheet_inserts[1],
                       'n_serum': labsheet_inserts[2],
                       'PBMC': labsheet_inserts[3],
                       'heparin': labsheet_inserts[4],
                       'header': labsheet_inserts[5]}
            doc.render(context)
            doc.save('#insert file path here' + 'checklist_' + str(ID) + '_' + name + '_' + vorname + '.docx')

            print('Studienmappe generiert: ' + vorname + ' ' + name + ' ' + str(ID))
            sheet1['J'+str(counter)] = 1

    counter += 1

wb_label_template.save('output/labels-prepared.xlsx')
wb_obj.save('')#insert file path here

print('Finished at row: ' + str(counter))
