# pip install openpyxl
# pip install python-docx
# pip install docxtpl

import openpyxl
from openpyxl.styles import PatternFill
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm
import datetime

# Defining file paths
admin_file = '' #insert file path here
#admin_file = 'admintestfile.xlsx'

# Use this to save the excel with formulas
wb_obj = openpyxl.load_workbook(admin_file)
sheet1 = wb_obj['COVID19_ZH_V5_Total']
sheet2 = wb_obj['counter']

# Use these to remove formulas (read-only)
wb_obj_f = openpyxl.load_workbook(admin_file, data_only=True)
sheet1_f = wb_obj_f['COVID19_ZH_V5_Total']
sheet2_f = wb_obj_f['counter']

# Opening label template
wb_label_template = openpyxl.load_workbook('dependencies/labels_template_phase5.xlsx')
label_sheet = wb_label_template.active

# Accessing the template checklist file
doc = DocxTemplate('dependencies/template checklist.docx')

# Defining group values
total_group1 = sheet1_f['H3'].value
total_group2 = sheet1_f['I3'].value

# Defining some static strings and images to be inserted into the checklist later on
header_g1 = InlineImage(doc, 'dependencies/g1_header.png', width=Mm(170))
header_g2 = InlineImage(doc, 'dependencies/g2_header.png', width=Mm(170))
header_g3 = InlineImage(doc, 'dependencies/g3_header.png', width=Mm(170))

small_g2 = InlineImage(doc, 'dependencies/g2_small.png', height=Mm(5))
small_g3 = InlineImage(doc, 'dependencies/g3_small.png', height=Mm(5))

tubes_g1 = '□ 1x Serum (8.5ml)\n' \
           '□ 1x Heparin (10ml)\n' \
           '□ 1x Heparin (2ml)'

tubes_g2 = '□ 1x Serum (8.5ml)\n' \
           '□ 1x Serum (8.5ml)\n' \
           '□ 1x Heparin (2ml)'

tubes_g3 = '□ 1x Serum (8.5ml)\n' \
           '□ 1x Serum (8.5ml)'

n_serum_g1 = '1x'

n_serum_g2_g3 = '2x'

PBMC_g1 = '□ Extraktion & Aliquotierung PBMCs - Heparin (10ml)\t\t\t\tTotal aliquots: 1\n' \
          '\t\t1 mL Vial(frozen -> ELISpot'

heparin_g1_g2 = '□ Handling Heparin (2ml)\n' \
                '\t\t1 T-Cell reactivity test'

# Counts how many participants of group 2 are scheduled for each test day (max 24 per day)
test_days_counter = {'01.03.2022': ['E1', sheet2['E1'].value],
                     '02.03.2022': ['E2', sheet2['E2'].value],
                     '03.03.2022': ['E3', sheet2['E3'].value],
                     '04.03.2022': ['E4', sheet2['E4'].value],
                     '05.03.2022': ['E5', sheet2['E5'].value],

                     '07.03.2022': ['E7', sheet2['E7'].value],
                     '08.03.2022': ['E8', sheet2['E8'].value],
                     '09.03.2022': ['E9', sheet2['E9'].value],
                     '10.03.2022': ['E10', sheet2['E10'].value],
                     '11.03.2022': ['E11', sheet2['E11'].value],
                     '12.03.2022': ['E12', sheet2['E12'].value],

                     '14.03.2022': ['E14', sheet2['E14'].value],
                     '15.03.2022': ['E15', sheet2['E15'].value],
                     '16.03.2022': ['E16', sheet2['E16'].value],
                     '17.03.2022': ['E17', sheet2['E17'].value],
                     '18.03.2022': ['E18', sheet2['E18'].value],

                     '21.03.2022': ['E21', sheet2['E21'].value],
                     '22.03.2022': ['E22', sheet2['E22'].value],
                     '23.03.2022': ['E23', sheet2['E23'].value],
                     '24.03.2022': ['E24', sheet2['E24'].value],
                     '25.03.2022': ['E25', sheet2['E25'].value],

                     '28.03.2022': ['E28', sheet2['E28'].value],
                     '29.03.2022': ['E29', sheet2['E29'].value],
                     '30.03.2022': ['E30', sheet2['E30'].value],
                     '31.03.2022': ['E31', sheet2['E31'].value]}

# Allocates a participant to group 1, 2, or 3 according to criteria
def group_allocater():

    global total_group1
    global total_group2
    global counter

    # Saturdays automatically Group 3
    if termin == '05.03.2022' or termin == '12.03.2022':
        sheet1['J' + str(counter)].value = 1
        sheet1['J' + str(counter)].fill = PatternFill(fgColor="FCE4D6", fill_type="solid")
        label_generator_g3()

        return [small_g3, tubes_g3, n_serum_g2_g3, '', '', header_g3]

    if total_group1 < 50:
        if termin == '01.03.2022' or termin == '02.03.2022' or termin == '03.03.2022':
            # Group 1
            total_group1 += 1
            sheet1['H' + str(counter)].value = 1
            sheet1['H' + str(counter)].fill = PatternFill(fgColor="DDEBF7", fill_type="solid")

            label_generator_g1()

            return ['Gruppe 1', tubes_g1, n_serum_g1, PBMC_g1, heparin_g1_g2, header_g1]

    else:
        if total_group2 < 350 and test_days_counter[termin][1] < 24:
            # Group 2
            test_days_counter[termin][1] += 1
            total_group2 += 1
            sheet2[test_days_counter[termin][0]] = test_days_counter[termin][1]
            sheet1['I' + str(counter)].value = 1
            sheet1['I' + str(counter)].fill = PatternFill(fgColor="FFF2CC", fill_type="solid")

            label_generator_g2()

            return [small_g2, tubes_g2, n_serum_g2_g3, '', heparin_g1_g2, header_g2]

        else:
            # Group 3
            sheet1['J' + str(counter)].value = 1
            sheet1['J' + str(counter)].fill = PatternFill(fgColor="FCE4D6", fill_type="solid")
            label_generator_g3()

            return [small_g3, tubes_g3, n_serum_g2_g3, '', '', header_g3]


def label_generator_g1():

    global lc

    label_ID = 'CI-T1-' + str(ID) + '-' + birthdate[-2:]

    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = ''
    label_sheet['C' + str(lc)] = termin
    lc += 1
    for i in range(4):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'BLOOD'
        label_sheet['C' + str(lc)] = termin
        lc += 1
    for i in range(6):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'SERUM'
        label_sheet['C' + str(lc)] = termin
        lc += 1
    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = 'PLASMA'
    label_sheet['C' + str(lc)] = termin
    lc += 1
    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = 'PBMC'
    label_sheet['C' + str(lc)] = termin
    lc += 1
    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = 'TCRT'
    label_sheet['C' + str(lc)] = termin
    lc += 1


# Generating labels for group 2
def label_generator_g2():

    global lc

    label_ID = 'CI-T1-' + str(ID) + '-' + birthdate[-2:]

    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = ''
    label_sheet['C' + str(lc)] = termin
    lc += 1
    for i in range(4):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'BLOOD'
        label_sheet['C' + str(lc)] = termin
        lc += 1
    for i in range(6):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'SERUM'
        label_sheet['C' + str(lc)] = termin
        lc += 1
    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = 'TCRT'
    label_sheet['C' + str(lc)] = termin
    lc += 1


# Generating labels for group 3
def label_generator_g3():

    global lc

    label_ID = 'CI-T1-' + str(ID) + '-' + birthdate[-2:]

    label_sheet['A' + str(lc)] = label_ID
    label_sheet['B' + str(lc)] = ''
    label_sheet['C' + str(lc)] = termin
    lc += 1
    for i in range(3):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'BLOOD'
        label_sheet['C' + str(lc)] = termin
        lc += 1
    for i in range(6):
        label_sheet['A' + str(lc)] = label_ID
        label_sheet['B' + str(lc)] = 'SERUM'
        label_sheet['C' + str(lc)] = termin
        lc += 1


# Iterating through the admin file. lc stands for label counter
counter = 11
lc = 2
for row in sheet1:

    # Extracting the data needed cell by cell
    ID = sheet1['D' + str(counter)].value
    vorname = sheet1['P'+str(counter)].value
    name = sheet1['Q'+str(counter)].value
    sex = sheet1['U'+str(counter)].value
    if sex == 1:
        sex = 'male'
    else:
        sex = 'female'

    try:
        birthdate = datetime.datetime.strptime(str(sheet1['R'+str(counter)].value.date()), "%Y-%m-%d").strftime("%d.%m.%Y")
        termin = datetime.datetime.strptime(str(sheet1['E' + str(counter)].value.date()), "%Y-%m-%d").strftime("%d.%m.%Y")
        time = sheet1['F' + str(counter)].value.strftime("%H:%M")
    except AttributeError:
        termin = 'ellis'

    fragebogenCode = sheet1['G' + str(counter)].value
    redCapFallErstellt = sheet1['B' + str(counter)].value
    studienMappeGeneriert = sheet1['C' + str(counter)].value

    # Pick tomorrow's appointments only or Mondays as well
    #if termin == (datetime.date.today() + datetime.timedelta(days=1)).strftime("%d.%m.%Y") or termin == (datetime.date.today() + datetime.timedelta(days=3)).strftime("%d.%m.%Y"):

    # Pick today's appointments only
    if termin == datetime.date.today().strftime("%d.%m.%Y"):

        # Pick relevant participants and generate the individual documents for them, update the admin file
        if redCapFallErstellt == 1 and studienMappeGeneriert is None:
            labsheet_inserts = group_allocater()
            context = {'ID': ID,
                       'vorname': vorname,
                       'name': name,
                       'sex': sex,
                       'birthdate': birthdate,
                       'termin': termin,
                       'time': time,
                       'fragebogenCode': fragebogenCode,
                       'label_ID': 'CI-T1-' + str(ID) + '-' + birthdate[-2:],
                       'group': labsheet_inserts[0],
                       'tubes': labsheet_inserts[1],
                       'n_serum': labsheet_inserts[2],
                       'PBMC': labsheet_inserts[3],
                       'heparin': labsheet_inserts[4],
                       'header': labsheet_inserts[5]}
            doc.render(context)
            doc.save('#insert file path here' + 'checklist_' + str(ID) + '_' + name + '_' + vorname + '.docx')

            print('Studienmappe generiert: ' + vorname + ' ' + name + ' ' + str(ID))
            sheet1['C'+str(counter)] = 1

    counter += 1

wb_label_template.save('output/labels-prepared.xlsx')
wb_obj.save('') #insert file path here
#wb_obj.save('admintestfile.xlsx')

print('Finished at row: ' + str(counter))
